"""Model sağlayıcı geçişleri, çoklu model çıktı yalıtımı ve Excel/CSV denetim testleri."""
from __future__ import annotations

import csv
import io
import json
import tempfile
import unittest
from pathlib import Path
import openpyxl

from terim_etmeni.config import Settings
from terim_etmeni.provider_store import ProviderConfig, ProviderConfigStore
from terim_etmeni.reporting import report_rows, write_reports
from terim_etmeni.service import AnalysisService
from terim_etmeni.web_app import Handler, Server


class ProviderTransitionAndExcelTests(unittest.TestCase):
    def setUp(self):
        self.tmp_dir = tempfile.TemporaryDirectory()
        self.root = Path(self.tmp_dir.name)
        self.dict_file = self.root / "dictionary.json"
        self.dict_file.write_text(
            json.dumps(
                {
                    "metadata": {"version": "2026-08-15"},
                    "terms": [
                        {"en": "attention mechanism", "tr": "dikkat mekanizması"},
                        {"en": "neural network", "tr": "yapay sinir ağı"},
                        {"en": "transaction", "tr": "işlem"},
                    ],
                }
            ),
            encoding="utf-8",
        )
        self.abbr_file = self.root / "abbreviations.json"
        self.abbr_file.write_text(
            json.dumps(
                {
                    "metadata": {"version": "2025-03-17"},
                    "abbreviations": [
                        {
                            "abbreviation": "LLM",
                            "expansion": "Large Language Model",
                            "turkish": "büyük dil modeli",
                        },
                        {
                            "abbreviation": "GPU",
                            "expansion": "Graphics Processing Unit",
                            "turkish": "grafik işlem birimi",
                        },
                    ],
                }
            ),
            encoding="utf-8",
        )
        self.output_dir = self.root / "output"
        self.output_dir.mkdir(parents=True, exist_ok=True)
        self.provider_file = self.root / "data" / "runtime" / "provider.json"
        self.provider_file.parent.mkdir(parents=True, exist_ok=True)

        self.settings = Settings(
            bootstrap_dictionary=self.dict_file,
            bootstrap_abbreviations=self.abbr_file,
            dictionary_state_dir=self.root / "state",
            output_dir=self.output_dir,
            provider_config_file=self.provider_file,
        )
        self.service = AnalysisService(self.settings)

    def tearDown(self):
        self.tmp_dir.cleanup()

    def _make_handler(self, path: str, method: str = "GET", headers: dict[str, str] | None = None, body: bytes = b""):
        handler = object.__new__(Handler)
        handler.path = path
        handler.command = method
        handler.request_version = "HTTP/1.1"
        headers_dict = dict(headers or {})
        if body and "Content-Length" not in headers_dict:
            headers_dict["Content-Length"] = str(len(body))

        from email.message import Message
        msg = Message()
        for k, v in headers_dict.items():
            msg[k] = v
        handler.headers = msg
        handler.rfile = io.BytesIO(body)
        handler.wfile = io.BytesIO()

        statuses = []
        response_headers = {}

        def send_response(code, message=None):
            statuses.append(code)

        def send_header(keyword, value):
            response_headers[keyword] = value

        def send_error(code, message=None, explain=None):
            statuses.append(code)

        handler.send_response = send_response
        handler.send_header = send_header
        handler.end_headers = lambda: None
        handler.send_error = send_error

        server = object.__new__(Server)
        server.service = self.service
        handler.server = server
        return handler, statuses, response_headers

    def test_provider_switching_deepseek_to_gemini_and_back(self):
        """1. Web arayüzü ve API üzerinden DeepSeek <-> Gemini 2.5 Flash geçişini test eder."""
        # 1.1 Başlangıç: DeepSeek ayarlarını POST /settings/save ile kaydet
        deepseek_body = b"provider=deepseek&model=deepseek-chat&api_key=sk-deepseek-test-key-12345&base_url="
        handler, statuses, headers = self._make_handler(
            "/settings/save",
            method="POST",
            headers={"Content-Type": "application/x-www-form-urlencoded"},
            body=deepseek_body,
        )
        handler.do_POST()
        self.assertEqual(statuses, [200])

        # provider.json dosyasının doğrulanması
        self.assertTrue(self.provider_file.is_file())
        saved_data = json.loads(self.provider_file.read_text(encoding="utf-8"))
        self.assertEqual(saved_data["provider"], "deepseek")
        self.assertEqual(saved_data["model"], "deepseek-chat")
        self.assertEqual(saved_data["api_key"], "sk-deepseek-test-key-12345")

        # Service provider status doğrulaması
        status = self.service.provider_status()
        self.assertEqual(status["provider"], "deepseek")
        self.assertEqual(status["model"], "deepseek-chat")
        self.assertTrue(status["has_key"])
        self.assertEqual(status["base_url"], "https://api.deepseek.com")
        self.assertTrue(self.service.using_api())
        models, err = self.service.installed_models()
        self.assertEqual(models, ["deepseek-chat"])
        self.assertEqual(err, "")

        # 1.2 Geçiş: Gemini 2.5 Flash'a geçiş (POST /api/settings)
        gemini_body = b"provider=google&model=gemini-2.5-flash&api_key=AIzaSyTestGoogleGeminiKey98765&base_url="
        handler2, statuses2, headers2 = self._make_handler(
            "/api/settings",
            method="POST",
            headers={"Content-Type": "application/x-www-form-urlencoded"},
            body=gemini_body,
        )
        handler2.do_POST()
        self.assertEqual(statuses2, [200])

        # provider.json dosyasının güncellendiğinin doğrulanması
        saved_gemini = json.loads(self.provider_file.read_text(encoding="utf-8"))
        self.assertEqual(saved_gemini["provider"], "google")
        self.assertEqual(saved_gemini["model"], "gemini-2.5-flash")
        self.assertEqual(saved_gemini["api_key"], "AIzaSyTestGoogleGeminiKey98765")

        # Service provider status doğrulaması
        status_gemini = self.service.provider_status()
        self.assertEqual(status_gemini["provider"], "google")
        self.assertEqual(status_gemini["model"], "gemini-2.5-flash")
        self.assertTrue(status_gemini["has_key"])
        self.assertEqual(status_gemini["base_url"], "https://generativelanguage.googleapis.com")
        models_gemini, err_gemini = self.service.installed_models()
        self.assertEqual(models_gemini, ["gemini-2.5-flash"])
        self.assertEqual(err_gemini, "")

        # 1.3 Anahtar koruma testi: Formda api_key boş gönderildiğinde mevcut anahtarın korunması
        switch_back_body = b"provider=deepseek&model=deepseek-chat&api_key=&base_url="
        handler3, statuses3, _ = self._make_handler(
            "/settings/save",
            method="POST",
            headers={"Content-Type": "application/x-www-form-urlencoded"},
            body=switch_back_body,
        )
        handler3.do_POST()
        self.assertEqual(statuses3, [200])
        saved_preserved = json.loads(self.provider_file.read_text(encoding="utf-8"))
        self.assertEqual(saved_preserved["provider"], "deepseek")
        self.assertEqual(saved_preserved["model"], "deepseek-chat")
        # Eski anahtar korunur
        self.assertEqual(saved_preserved["api_key"], "AIzaSyTestGoogleGeminiKey98765")

    def test_output_directory_isolation_between_deepseek_and_gemini(self):
        """Her iki modelin ürettiği raporların output/ altında birbirini ezmeden saklandığını doğrular."""
        deepseek_result = {
            "document": "academic_paper.pdf",
            "model": "deepseek-chat",
            "dictionary_version": "2026-08-15",
            "analysis_status": "complete",
            "missing_terms": [
                {
                    "term": "low-rank adaptation",
                    "context": "LoRA uses low-rank adaptation matrices.",
                    "pages": [1, 2],
                    "occurrence_count": 5,
                    "review_priority": "high",
                }
            ],
            "dictionary_matches": [
                {
                    "term": "neural network",
                    "translations": ["yapay sinir ağı"],
                    "match_type": "exact",
                    "context": "Trained on a deep neural network.",
                    "pages": [1],
                    "occurrence_count": 3,
                }
            ],
            "possible_matches": [],
            "rejected_candidates": [],
        }

        gemini_result = {
            "document": "academic_paper.pdf",
            "model": "gemini-2.5-flash",
            "dictionary_version": "2026-08-15",
            "analysis_status": "complete",
            "missing_terms": [
                {
                    "term": "mixture of experts",
                    "context": "MoE routes tokens to sparse mixture of experts layers.",
                    "pages": [3, 4],
                    "occurrence_count": 8,
                    "review_priority": "high",
                },
                {
                    "term": "flash attention",
                    "context": "Optimized memory I/O using flash attention kernels.",
                    "pages": [2],
                    "occurrence_count": 4,
                    "review_priority": "medium",
                },
            ],
            "dictionary_matches": [
                {
                    "term": "attention mechanism",
                    "translations": ["dikkat mekanizması"],
                    "match_type": "exact",
                    "context": "Self-attention mechanism captures long-range dependencies.",
                    "pages": [1, 2],
                    "occurrence_count": 12,
                }
            ],
            "possible_matches": [
                {
                    "term": "LLM",
                    "possible_dictionary_terms": [
                        {"en": "Large Language Model", "tr": "büyük dil modeli"}
                    ],
                    "context": "LLM inference requires high memory bandwidth.",
                    "pages": [5],
                    "occurrence_count": 3,
                }
            ],
            "rejected_candidates": [],
        }

        # 1. DeepSeek raporlarını yaz
        ds_json, ds_csv = write_reports(deepseek_result, self.output_dir)
        ds_xlsx = ds_json.parent / "academic_paper_terim_raporu.xlsx"

        # 2. Gemini 2.5 Flash raporlarını yaz
        gem_json, gem_csv = write_reports(gemini_result, self.output_dir)
        gem_xlsx = gem_json.parent / "academic_paper_terim_raporu.xlsx"

        # Dosya konumları ve klasör yalıtımı kontrolü
        self.assertEqual(ds_json.parent.name, "academic_paper")
        self.assertEqual(ds_json.parent.parent.name, "deepseek-chat")
        self.assertEqual(gem_json.parent.name, "academic_paper")
        self.assertEqual(gem_json.parent.parent.name, "gemini-2.5-flash")

        # Her iki model klasöründeki dosyaların varlığı ve bağımsızlığı
        self.assertTrue(ds_json.is_file())
        self.assertTrue(ds_csv.is_file())
        self.assertTrue(ds_xlsx.is_file())

        self.assertTrue(gem_json.is_file())
        self.assertTrue(gem_csv.is_file())
        self.assertTrue(gem_xlsx.is_file())

        # İçeriklerin birbirini ezmediğinin ve doğru modellere ait olduğunun doğrulanması
        ds_json_content = json.loads(ds_json.read_text(encoding="utf-8"))
        self.assertEqual(ds_json_content["model"], "deepseek-chat")
        self.assertEqual(len(ds_json_content["missing_terms"]), 1)
        self.assertEqual(ds_json_content["missing_terms"][0]["term"], "low-rank adaptation")

        gem_json_content = json.loads(gem_json.read_text(encoding="utf-8"))
        self.assertEqual(gem_json_content["model"], "gemini-2.5-flash")
        self.assertEqual(len(gem_json_content["missing_terms"]), 2)
        self.assertEqual(gem_json_content["missing_terms"][0]["term"], "mixture of experts")

        # Web /reports indirme rotalarının her iki klasör için sorunsuz çalışması
        handler_ds, statuses_ds, headers_ds = self._make_handler(
            "/reports/deepseek-chat%2Facademic_paper%2Facademic_paper_terim_raporu.xlsx"
        )
        handler_ds.do_GET()
        self.assertEqual(statuses_ds, [200])
        self.assertEqual(headers_ds["Content-Type"], "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

        handler_gem, statuses_gem, headers_gem = self._make_handler(
            "/reports/gemini-2.5-flash%2Facademic_paper%2Facademic_paper_terim_raporu.xlsx"
        )
        handler_gem.do_GET()
        self.assertEqual(statuses_gem, [200])
        self.assertEqual(headers_gem["Content-Type"], "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    def test_gemini_excel_sheets_turkish_chars_empty_committee_and_csv_sync(self):
        """2. Gemini Excel (.xlsx) dosyasının sekmelerini, Türkçe karakterlerini, boş komite sütunlarını ve CSV uyumunu openpyxl ile test eder."""
        turkish_context = "Örnek Türkçe metin: Çözümleme, genişletilmiş işlem, ölçüt ve yöney uzayı; şifreleme & özet."
        gemini_result = {
            "document": "gemini_denetim_belgesi.pdf",
            "model": "gemini-2.5-flash",
            "dictionary_version": "2026-08-15",
            "analysis_status": "complete",
            "missing_terms": [
                {
                    "term": "sparse mixture of experts",
                    "context": "Sparse mixture of experts (SMoE) reduces compute during routing.",
                    "pages": [1, 3],
                    "occurrence_count": 6,
                    "review_priority": "high",
                },
                {
                    "term": "speculative decoding",
                    "context": turkish_context,
                    "pages": [2],
                    "occurrence_count": 2,
                    "review_priority": "medium",
                },
                {
                    "term": "düşük_öncelikli_kavram [v1.0]",
                    "context": "Açıklama cümlesi: â, î, û, ç, ğ, ı, ö, ş, ü karakterleri.",
                    "pages": [5],
                    "occurrence_count": 1,
                    "review_priority": "low",
                },
            ],
            "dictionary_matches": [
                {
                    "term": "transactions",
                    "translations": ["işlem"],
                    "match_type": "singular_variant",
                    "context": "All financial transactions are verified.",
                    "pages": [1, 2, 4],
                    "occurrence_count": 15,
                },
                {
                    "term": "attention mechanism",
                    "translations": ["dikkat mekanizması"],
                    "match_type": "exact",
                    "context": "Multi-head attention mechanism architecture.",
                    "pages": [2, 3],
                    "occurrence_count": 8,
                },
            ],
            "possible_matches": [
                {
                    "term": "GPU",
                    "possible_dictionary_terms": [
                        {"en": "Graphics Processing Unit", "tr": "grafik işlem birimi"}
                    ],
                    "context": "Workload is distributed across 8 GPU nodes.",
                    "pages": [4],
                    "occurrence_count": 4,
                }
            ],
            "rejected_candidates": [],
        }

        json_p, csv_p = write_reports(gemini_result, self.output_dir)
        xlsx_p = json_p.parent / "gemini_denetim_belgesi_terim_raporu.xlsx"

        self.assertTrue(xlsx_p.is_file())
        self.assertTrue(csv_p.is_file())

        # ====================================================
        # OPENPYXL İLE EXCEL (.XLSX) DENETİMİ
        # ====================================================
        wb = openpyxl.load_workbook(xlsx_p)

        # 2.1 Sekme Adları Doğrulaması
        expected_sheets = ["Eksik Terimler (İnceleme)", "Sözlükte Bulunanlar"]
        self.assertEqual(wb.sheetnames, expected_sheets)

        # ----------------------------------------------------
        # SEKME 1: Eksik Terimler (İnceleme)
        # ----------------------------------------------------
        ws1 = wb["Eksik Terimler (İnceleme)"]
        self.assertTrue(ws1.views.sheetView[0].showGridLines)
        self.assertEqual(ws1.freeze_panes, "A5")

        # Başlık ve Alt Başlık Bannerları
        self.assertIn("TBD BİLİŞİM SÖZLÜĞÜ — EKSİK TERİM İNCELEME LİSTESİ", str(ws1["A1"].value))
        self.assertIn("gemini_denetim_belgesi.pdf", str(ws1["A2"].value))
        self.assertIn("gemini-2.5-flash", str(ws1["A2"].value))
        self.assertIn("3 Adet", str(ws1["A2"].value))

        # Tablo Başlıkları (Satır 4)
        expected_ws1_headers = [
            "No",
            "İngilizce Terim",
            "Makaledeki Örnek Cümle (Bağlam)",
            "Önerilen Türkçe Karşılık (Komite)",
            "Öncelik",
            "Geçiş",
            "Sayfalar",
            "Komite Notu / Karar",
        ]
        for col_i, header_text in enumerate(expected_ws1_headers, 1):
            cell_val = ws1.cell(row=4, column=col_i).value
            self.assertEqual(cell_val, header_text)

        # Veri Satırları Denetimi (Satır 5, 6, 7)
        # Satır 5: "sparse mixture of experts"
        self.assertEqual(ws1.cell(row=5, column=1).value, 1)
        self.assertEqual(ws1.cell(row=5, column=2).value, "sparse mixture of experts")
        self.assertIn("Sparse mixture of experts", str(ws1.cell(row=5, column=3).value))
        # 2.2 BOŞ KOMİTE SÜTUNLARI KONTROLÜ
        self.assertIn(ws1.cell(row=5, column=4).value, ("", None))  # Komite Karşılık Sütunu boş olmalı
        self.assertIn(ws1.cell(row=5, column=8).value, ("", None))  # Komite Notu Sütunu boş olmalı
        self.assertEqual(ws1.cell(row=5, column=5).value, "Yüksek")  # Öncelik
        self.assertEqual(ws1.cell(row=5, column=6).value, 6)  # Geçiş
        self.assertEqual(ws1.cell(row=5, column=7).value, "1, 3")  # Sayfalar

        # Satır 6: "speculative decoding" ve Türkçe Karakter Kontrolü
        self.assertEqual(ws1.cell(row=6, column=2).value, "speculative decoding")
        ctx_row6 = str(ws1.cell(row=6, column=3).value)
        self.assertIn("Çözümleme", ctx_row6)
        self.assertIn("işlem", ctx_row6)
        self.assertIn("ölçüt", ctx_row6)
        self.assertIn("yöney", ctx_row6)
        self.assertIn("şifreleme", ctx_row6)
        self.assertIn(ws1.cell(row=6, column=4).value, ("", None))  # Boş komite sütunu
        self.assertIn(ws1.cell(row=6, column=8).value, ("", None))  # Boş komite notu
        self.assertEqual(ws1.cell(row=6, column=5).value, "Orta")

        # Satır 7: "düşük_öncelikli_kavram [v1.0]" ve şapkalı karakterler
        self.assertEqual(ws1.cell(row=7, column=2).value, "düşük_öncelikli_kavram [v1.0]")
        ctx_row7 = str(ws1.cell(row=7, column=3).value)
        self.assertIn("â, î, û, ç, ğ, ı, ö, ş, ü", ctx_row7)
        self.assertIn(ws1.cell(row=7, column=4).value, ("", None))  # Boş komite sütunu
        self.assertIn(ws1.cell(row=7, column=8).value, ("", None))  # Boş komite notu
        self.assertEqual(ws1.cell(row=7, column=5).value, "Düşük")

        # Sütun 4 Renk Biçimlendirmesi (Giriş Kutusu Vurgusu - FFFDF0)
        for r in (5, 6, 7):
            cell_prop = ws1.cell(row=r, column=4)
            self.assertEqual(cell_prop.fill.start_color.rgb, "00FFFDF0")

        # ----------------------------------------------------
        # SEKME 2: Sözlükte Bulunanlar
        # ----------------------------------------------------
        ws2 = wb["Sözlükte Bulunanlar"]
        self.assertTrue(ws2.views.sheetView[0].showGridLines)
        self.assertEqual(ws2.freeze_panes, "A5")

        # Başlık ve Alt Başlık Bannerları
        self.assertIn("TBD BİLİŞİM SÖZLÜĞÜNDE BULUNAN TERİMLER VE KISALTMALAR", str(ws2["A1"].value))
        self.assertIn("Toplam Eşleşen Terim: 3 Adet", str(ws2["A2"].value))
        self.assertIn("2026-08-15", str(ws2["A2"].value))

        # Tablo Başlıkları (Satır 4)
        expected_ws2_headers = [
            "No",
            "İngilizce Terim",
            "TBD Sözlük Türkçe Karşılığı",
            "Eşleşme Durumu",
            "Makaledeki Örnek Cümle (Bağlam)",
            "Geçiş",
            "Sayfalar",
        ]
        for col_i, header_text in enumerate(expected_ws2_headers, 1):
            cell_val = ws2.cell(row=4, column=col_i).value
            self.assertEqual(cell_val, header_text)

        # Bulunan terimler: transactions, attention mechanism, GPU
        # Satır 5: transactions (Çoğul Eşleşme)
        self.assertEqual(ws2.cell(row=5, column=2).value, "transactions")
        self.assertEqual(ws2.cell(row=5, column=3).value, "işlem")
        self.assertEqual(ws2.cell(row=5, column=4).value, "Sözlükte Bulundu (Çoğul Eşleşme)")
        self.assertEqual(ws2.cell(row=5, column=6).value, 15)
        self.assertEqual(ws2.cell(row=5, column=7).value, "1, 2, 4")

        # Satır 6: attention mechanism (Sözlükte Kayıtlı)
        self.assertEqual(ws2.cell(row=6, column=2).value, "attention mechanism")
        self.assertEqual(ws2.cell(row=6, column=3).value, "dikkat mekanizması")
        self.assertEqual(ws2.cell(row=6, column=4).value, "Sözlükte Kayıtlı")

        # Satır 7: GPU (TBD Kısaltması)
        self.assertEqual(ws2.cell(row=7, column=2).value, "GPU")
        self.assertEqual(ws2.cell(row=7, column=3).value, "grafik işlem birimi")
        self.assertEqual(ws2.cell(row=7, column=4).value, "TBD Kısaltması")

        # ====================================================
        # CSV (.CSV) UYUMU VE SENKRONİZASYON DENETİMİ
        # ====================================================
        csv_raw = csv_p.read_bytes()
        # UTF-8-SIG (BOM = \xef\xbb\xbf) Kontrolü
        self.assertTrue(csv_raw.startswith(b"\xef\xbb\xbf"), "CSV dosyası Excel uyumlu UTF-8-SIG (BOM) ile başlamalı.")

        csv_text = csv_raw.decode("utf-8-sig")
        csv_reader = csv.DictReader(io.StringIO(csv_text), delimiter=";")

        rows = list(csv_reader)
        # Toplam satır sayısı = 3 eksik + 2 sözlükte bulunan + 1 kısaltma = 6 satır
        self.assertEqual(len(rows), 6)

        # CSV Başlıkları
        expected_csv_fields = [
            "İnceleme Durumu",
            "Eşleşme Türü",
            "Öncelik",
            "Önerilen İşlem",
            "İngilizce Terim",
            "Türkçe Karşılık",
            "Makaledeki Bağlam (Örnek Cümle)",
            "Önerilen Türkçe Karşılık (Komite)",
            "Yakın Sözlük Eşleşmesi",
            "Kanıt Sayfaları",
            "PDF'deki Geçiş Sayısı",
            "Açıklama",
        ]
        self.assertEqual(csv_reader.fieldnames, expected_csv_fields)

        # Eksik terimlerin CSV'deki boş komite sütunları ve Türkçe karakter kontrolü
        missing_csv = [r for r in rows if r["İnceleme Durumu"] == "İnceleme gerekli"]
        self.assertEqual(len(missing_csv), 3)

        # 1. Eksik Terim
        self.assertEqual(missing_csv[0]["İngilizce Terim"], "sparse mixture of experts")
        self.assertEqual(missing_csv[0]["Önerilen Türkçe Karşılık (Komite)"], "")
        self.assertEqual(missing_csv[0]["Öncelik"], "Yüksek")
        self.assertEqual(missing_csv[0]["Eşleşme Türü"], "Sözlükte Yok (Eksik Terim)")

        # 2. Eksik Terim (Türkçe Karakterler)
        self.assertEqual(missing_csv[1]["İngilizce Terim"], "speculative decoding")
        self.assertEqual(missing_csv[1]["Önerilen Türkçe Karşılık (Komite)"], "")
        self.assertIn("Çözümleme", missing_csv[1]["Makaledeki Bağlam (Örnek Cümle)"])
        self.assertIn("şifreleme", missing_csv[1]["Makaledeki Bağlam (Örnek Cümle)"])

        # 3. Eksik Terim
        self.assertEqual(missing_csv[2]["İngilizce Terim"], "düşük_öncelikli_kavram [v1.0]")
        self.assertEqual(missing_csv[2]["Önerilen Türkçe Karşılık (Komite)"], "")
        self.assertIn("â, î, û, ç, ğ, ı, ö, ş, ü", missing_csv[2]["Makaledeki Bağlam (Örnek Cümle)"])

        # Bulunan terimlerin CSV kontrolü
        found_csv = [r for r in rows if r["İnceleme Durumu"] == "Sözlükte bulundu"]
        self.assertEqual(len(found_csv), 2)
        self.assertEqual(found_csv[0]["İngilizce Terim"], "transactions")
        self.assertEqual(found_csv[0]["Türkçe Karşılık"], "işlem")
        self.assertEqual(found_csv[0]["Eşleşme Türü"], "Sözlükte Bulundu (Çoğul Eşleşme)")

        self.assertEqual(found_csv[1]["İngilizce Terim"], "attention mechanism")
        self.assertEqual(found_csv[1]["Türkçe Karşılık"], "dikkat mekanizması")
        self.assertEqual(found_csv[1]["Eşleşme Türü"], "Sözlükte Kayıtlı")

        # Kısaltma CSV kontrolü
        abbr_csv = [r for r in rows if r["İnceleme Durumu"] == "Yakın eşleşme"]
        self.assertEqual(len(abbr_csv), 1)
        self.assertEqual(abbr_csv[0]["İngilizce Terim"], "GPU")
        self.assertEqual(abbr_csv[0]["Eşleşme Türü"], "TBD Kısaltması")
        self.assertIn("Graphics Processing Unit → grafik işlem birimi", abbr_csv[0]["Yakın Sözlük Eşleşmesi"])


if __name__ == "__main__":
    unittest.main()
