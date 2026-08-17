from __future__ import annotations

import json
import tempfile
import unittest
from pathlib import Path

from terim_etmeni.reporting import _clean_str, report_rows, write_reports



class ReportingTests(unittest.TestCase):
    def test_write_reports_creates_two_sheet_xlsx_and_csv(self):
        result = {
            "document": "test_doc.pdf",
            "model": "test-model",
            "dictionary_version": "2026-07-20",
            "missing_terms": [
                {
                    "term": "soft state",
                    "context": "Nodes maintain soft state across rounds.",
                    "pages": [8],
                    "occurrence_count": 2,
                    "review_priority": "high",
                }
            ],
            "dictionary_matches": [
                {
                    "term": "distributed systems",
                    "translations": ["dağıtımlı dizge"],
                    "match_type": "singular_variant",
                    "context": "Distributed systems scale horizontally.",
                    "pages": [1, 2],
                    "occurrence_count": 10,
                }
            ],
            "possible_matches": [
                {
                    "term": "RBAC",
                    "possible_dictionary_terms": [
                        {"en": "Role-Based Access Control", "tr": "role dayalı erişim denetimi"}
                    ],
                    "context": "RBAC policy is enforced.",
                    "pages": [5],
                    "occurrence_count": 1,
                }
            ],
            "rejected_candidates": [],
        }

        with tempfile.TemporaryDirectory() as tmp_dir:
            json_path, csv_path = write_reports(result, Path(tmp_dir))
            xlsx_path = json_path.with_name("test_doc_terim_raporu.xlsx")

            self.assertTrue(json_path.is_file())
            self.assertTrue(csv_path.is_file())
            self.assertTrue(xlsx_path.is_file())

            # CSV içeriği kontrolü
            csv_content = csv_path.read_text(encoding="utf-8-sig")
            self.assertIn("soft state", csv_content)
            self.assertIn("Nodes maintain soft state across rounds.", csv_content)
            self.assertIn("Makaledeki Bağlam (Örnek Cümle)", csv_content)
            self.assertIn(";", csv_content)

            # XLSX sayfaları kontrolü
            import openpyxl
            wb = openpyxl.load_workbook(xlsx_path)
            self.assertEqual(wb.sheetnames, ["Eksik Terimler (İnceleme)", "Sözlükte Bulunanlar"])
            ws1 = wb["Eksik Terimler (İnceleme)"]
            self.assertEqual(ws1["B5"].value, "soft state")
            self.assertIn("soft state", str(ws1["C5"].value))

    def test_reporting_robustness_with_turkish_chars_long_text_and_special_symbols(self):
        turkish_text = "Özel Türkçe karakterler: ç, Ç, ğ, Ğ, ı, İ, ö, Ö, ş, Ş, ü, Ü, â, î, û."
        long_context = "Çok uzun metin bağlamı: " + ("Detaylı açıklama cümlesi; noktalı virgül, 'tırnak' ve \"çift tırnak\". " * 50)
        complex_term = "kriptografik-özet_fonksiyonu [SHA-256] <v2.0> & (özel = + - @ formül)"
        control_char_context = "Metin içinde geçersiz kontrol karakterleri: \x00, \x0b, \x0c, \x1f ve yeni satır \n ile tab \t."

        result = {
            "document": "türkçe_makale_örneği.pdf",
            "model": "gpt-4o-mini",
            "dictionary_version": "2026-08-15",
            "missing_terms": [
                {
                    "term": complex_term,
                    "context": long_context,
                    "pages": [1, 3, 5],
                    "occurrence_count": 7,
                    "review_priority": "high",
                },
                {
                    "term": "kontrol_karakter_terimi",
                    "context": control_char_context,
                    "pages": [],
                    "occurrence_count": 0,
                    "review_priority": "low",
                },
                {
                    "term": "boş_bağlam_terimi",
                    "context": "",
                    "pages": [12],
                    "occurrence_count": 1,
                    "review_priority": "medium",
                },
            ],
            "dictionary_matches": [
                {
                    "term": "yapay zekâ",
                    "translations": [turkish_text],
                    "match_type": "exact",
                    "context": turkish_text,
                    "pages": [1],
                    "occurrence_count": 12,
                }
            ],
            "possible_matches": [],
            "rejected_candidates": [],
        }

        with tempfile.TemporaryDirectory() as tmp_dir:
            json_path, csv_path = write_reports(result, Path(tmp_dir))
            xlsx_path = json_path.with_name("türkçe_makale_örneği_terim_raporu.xlsx")

            self.assertTrue(json_path.is_file())
            self.assertTrue(csv_path.is_file())
            self.assertTrue(xlsx_path.is_file())

            # CSV validation
            csv_content = csv_path.read_text(encoding="utf-8-sig")
            self.assertIn("ç, Ç, ğ, Ğ, ı, İ, ö, Ö, ş, Ş, ü, Ü", csv_content)
            self.assertIn("kriptografik-özet_fonksiyonu", csv_content)

            # Excel validation with openpyxl
            import openpyxl
            wb = openpyxl.load_workbook(xlsx_path)
            ws1 = wb["Eksik Terimler (İnceleme)"]
            self.assertIn("kriptografik-özet_fonksiyonu", ws1["B5"].value)
            self.assertEqual(ws1["B6"].value, "kontrol_karakter_terimi")
            self.assertEqual(ws1["B7"].value, "boş_bağlam_terimi")
            self.assertIn(ws1["C7"].value, ("", None))


            # Verify illegal control chars were removed from Excel cell
            cell_val = ws1["C6"].value
            self.assertNotIn("\x00", cell_val)
            self.assertNotIn("\x0b", cell_val)
            self.assertNotIn("\x0c", cell_val)
            self.assertNotIn("\x1f", cell_val)

            ws2 = wb["Sözlükte Bulunanlar"]
            self.assertIn("yapay zekâ", ws2["B5"].value)
            self.assertIn("Özel Türkçe karakterler", ws2["C5"].value)

    def test_clean_str_strips_control_characters(self):
        self.assertEqual(_clean_str(None), "")
        self.assertEqual(_clean_str("Hello\x00World\x0b\x0cTest\x1fEnd"), "HelloWorldTestEnd")
        self.assertEqual(_clean_str("Türkçe ğüşiöç"), "Türkçe ğüşiöç")


if __name__ == "__main__":
    unittest.main()

