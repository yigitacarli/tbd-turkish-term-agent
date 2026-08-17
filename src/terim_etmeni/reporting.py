"""Analiz sonuçlarını XLSX, CSV ve JSON biçimlerinde kaydetme modülü."""
from __future__ import annotations

import csv
import json
import re
import unicodedata
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    _HAS_OPENPYXL = True
except ImportError:
    _HAS_OPENPYXL = False

ILLEGAL_CHARACTERS_RE = re.compile(r"[\000-\010]|[\013-\014]|[\016-\037]")


def _clean_str(val: object) -> str:
    """Excel XML standartlarına aykırı kontrol karakterlerini temizler."""
    if val is None:
        return ""
    return ILLEGAL_CHARACTERS_RE.sub("", str(val))



STATUS_LABELS = {
    "missing_terms": "İnceleme gerekli",
    "possible_matches": "Yakın eşleşme",
    "dictionary_matches": "Sözlükte bulundu",
    "rejected_candidates": "Elenen aday",
}


def _pages(item: dict[str, object]) -> str:
    return ", ".join(str(page) for page in item.get("pages", []))


def report_rows(result: dict[str, object]) -> list[dict[str, object]]:
    """CSV ve veri aktarımı için zenginleştirilmiş satırlar üretir."""
    rows: list[dict[str, object]] = []
    for group in ("missing_terms", "possible_matches", "dictionary_matches", "rejected_candidates"):
        status = STATUS_LABELS[group]
        items = result.get(group, [])
        if not isinstance(items, list):
            continue
        for item in items:
            if not isinstance(item, dict):
                continue
            translations = item.get("translations", [])
            suggestions = item.get("possible_dictionary_terms", [])
            suggestion_text = " | ".join(
                "{} → {}".format(value.get("en", ""), value.get("tr", ""))
                for value in suggestions
                if isinstance(value, dict)
            ) if isinstance(suggestions, list) else ""

            match_type = str(item.get("match_type", ""))
            if match_type == "singular_variant":
                type_label = "Sözlükte Bulundu (Çoğul Eşleşme)"
            elif match_type == "exact":
                type_label = "Sözlükte Kayıtlı"
            elif group == "possible_matches":
                type_label = "TBD Kısaltması"
            else:
                type_label = "Sözlükte Yok (Eksik Terim)"

            if group == "missing_terms":
                priority = {
                    "high": "Yüksek",
                    "medium": "Orta",
                    "low": "Düşük",
                }.get(str(item.get("review_priority", "high")), "Yüksek")
                action = "Terimi insan incelemesine al"
                explanation = "TBD Bilişim Sözlüğü'nde bulunamadı; uzman incelemesi ve Türkçe karşılık önerisi bekleniyor."
            elif group == "possible_matches":
                priority = "Orta"
                action = "Kısaltma açılımını doğrula"
                explanation = "TBD Kısaltmalar tablosunda açılımı ve Türkçe karşılığı bulundu."
            elif group == "dictionary_matches":
                priority = "Bilgi"
                action = "İşlem gerekmez"
                explanation = "Sözlükte doğrudan Türkçe karşılık bulundu."
            else:
                priority = "Düşük"
                action = "Yok say"
                explanation = "Biçimsel gürültü veya düşük güvenli aday olarak elendi."

            rows.append(
                {
                    "İnceleme Durumu": status,
                    "Eşleşme Türü": type_label,
                    "Öncelik": priority,
                    "Önerilen İşlem": action,
                    "İngilizce Terim": item.get("term", ""),
                    "Türkçe Karşılık": ", ".join(str(value) for value in translations),
                    "Makaledeki Bağlam (Örnek Cümle)": item.get("context", ""),
                    "Önerilen Türkçe Karşılık (Komite)": "",
                    "Yakın Sözlük Eşleşmesi": suggestion_text,
                    "Kanıt Sayfaları": _pages(item),
                    "PDF'deki Geçiş Sayısı": item.get("occurrence_count", 0),
                    "Açıklama": explanation,
                }
            )
    return rows


def format_terminal_report(result: dict[str, object]) -> str:
    """Terminale kısa, karar odaklı bir özet yazar."""
    lines = ["", "=== TERİM RAPORU ===", "CSV ve Excel'de önce yüksek öncelikli inceleme adayları yer alır."]
    titles = {
        "missing_terms": "SÖZLÜKTE OLMAYANLAR (İNCELEME GEREKLİ)",
        "possible_matches": "TBD KISALTMALARI / YAKIN EŞLEŞMELER",
        "dictionary_matches": "SÖZLÜKTE BULUNANLAR",
    }
    for group, title in titles.items():
        items = result.get(group, [])
        values = items if isinstance(items, list) else []
        lines.extend(["", "{} ({})".format(title, len(values))])
        if not values:
            lines.append("  - Yok")
            continue
        for item in values[:10]:
            if not isinstance(item, dict):
                continue
            detail = ""
            if group == "dictionary_matches":
                translations = item.get("translations", [])
                if translations:
                    detail = " -> " + " | ".join(str(value) for value in translations)
            elif group == "possible_matches":
                matches = item.get("possible_dictionary_terms", [])
                detail = " ~ " + " | ".join(
                    "{} -> {}".format(value.get("en", ""), value.get("tr", ""))
                    for value in matches
                    if isinstance(value, dict)
                )
            pages = _pages(item)
            lines.append(
                "  - {}{} [sayfa: {}; geçiş: {}]".format(
                    item.get("term", ""),
                    detail,
                    pages or "-",
                    item.get("occurrence_count", 0),
                )
            )
        if len(values) > 10:
            lines.append("  - … {} aday daha; ayrıntılar Excel/CSV raporunda.".format(len(values) - 10))
    return "\n".join(lines)


def _export_styled_xlsx(
    result: dict[str, object],
    xlsx_path: Path,
) -> None:
    """İki sekmeli, yüksek kontrastlı ve insan incelemesine hazır Excel (.xlsx) raporu üretir."""
    if not _HAS_OPENPYXL:
        return

    wb = openpyxl.Workbook()
    # Varsayılan sayfayı temizle
    wb.remove(wb.active)

    doc_name = Path(str(result.get("document", "Dokuman"))).name
    model_name = str(result.get("model") or "Belirtilmedi")
    dict_version = str(result.get("dictionary_version") or "2026-07-20")

    missing = [item for item in result.get("missing_terms", []) if isinstance(item, dict)]
    found = [item for item in result.get("dictionary_matches", []) if isinstance(item, dict)]
    possible = [item for item in result.get("possible_matches", []) if isinstance(item, dict)]

    thin_border = Border(
        left=Side(style="thin", color="D1D5DB"),
        right=Side(style="thin", color="D1D5DB"),
        top=Side(style="thin", color="D1D5DB"),
        bottom=Side(style="thin", color="D1D5DB"),
    )
    input_box_border = Border(
        left=Side(style="medium", color="0284C7"),
        right=Side(style="medium", color="0284C7"),
        top=Side(style="medium", color="0284C7"),
        bottom=Side(style="medium", color="0284C7"),
    )
    alt_fill = PatternFill(start_color="F8FAFC", fill_type="solid")
    input_fill = PatternFill(start_color="FFFDF0", fill_type="solid")

    priority_styles = {
        "Yüksek": (PatternFill(start_color="FEE2E2", fill_type="solid"), Font(name="Calibri", size=10, bold=True, color="991B1B")),
        "Orta": (PatternFill(start_color="FEF3C7", fill_type="solid"), Font(name="Calibri", size=10, bold=True, color="92400E")),
        "Düşük": (PatternFill(start_color="E0F2FE", fill_type="solid"), Font(name="Calibri", size=10, color="075985")),
    }

    # ==========================================
    # SEKME 1: EKSİK TERİMLER (İNCELEME LİSTESİ)
    # ==========================================
    ws1 = wb.create_sheet(title="Eksik Terimler (İnceleme)")
    ws1.views.sheetView[0].showGridLines = True

    # Başlık Bannerı
    ws1.merge_cells("A1:H1")
    title1 = ws1["A1"]
    title1.value = _clean_str(f"🎯 TBD BİLİŞİM SÖZLÜĞÜ — EKSİK TERİM İNCELEME LİSTESİ")
    title1.font = Font(name="Calibri", size=14, bold=True, color="1E3A5F")
    title1.alignment = Alignment(vertical="center")
    ws1.row_dimensions[1].height = 28

    ws1.merge_cells("A2:H2")
    sub1 = ws1["A2"]
    sub1.value = _clean_str(f"İncelenen Belge: {doc_name} | Analiz Modeli: {model_name} | İncelenecek Eksik Terim: {len(missing)} Adet")
    sub1.font = Font(name="Calibri", size=10, italic=True, color="475569")
    sub1.alignment = Alignment(vertical="center")
    ws1.row_dimensions[2].height = 20

    header_row_1 = 4
    ws1.row_dimensions[header_row_1].height = 28
    headers_1 = [
        ("No", 6),
        ("İngilizce Terim", 26),
        ("Makaledeki Örnek Cümle (Bağlam)", 55),
        ("Önerilen Türkçe Karşılık (Komite)", 34),
        ("Öncelik", 14),
        ("Geçiş", 10),
        ("Sayfalar", 14),
        ("Komite Notu / Karar", 30),
    ]

    header_fill_1 = PatternFill(start_color="1E3A5F", fill_type="solid")
    header_font_1 = Font(name="Calibri", size=11, bold=True, color="FFFFFF")

    for col_idx, (h_text, col_w) in enumerate(headers_1, 1):
        cell = ws1.cell(row=header_row_1, column=col_idx, value=h_text)
        cell.fill = header_fill_1
        cell.font = header_font_1
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        col_letter = get_column_letter(col_idx)
        ws1.column_dimensions[col_letter].width = col_w

    for idx, item in enumerate(missing, 1):
        row_idx = header_row_1 + idx
        ws1.row_dimensions[row_idx].height = 36
        is_even = (idx % 2 == 0)

        c_no = ws1.cell(row=row_idx, column=1, value=idx)
        c_no.alignment = Alignment(horizontal="center", vertical="center")
        c_no.font = Font(name="Calibri", size=10, color="64748B")

        c_term = ws1.cell(row=row_idx, column=2, value=_clean_str(item.get("term", "")))
        c_term.font = Font(name="Calibri", size=11, bold=True, color="1E3A5F")
        c_term.alignment = Alignment(vertical="center")

        c_ctx = ws1.cell(row=row_idx, column=3, value=_clean_str(item.get("context", "")))
        c_ctx.font = Font(name="Calibri", size=10, italic=True, color="334155")
        c_ctx.alignment = Alignment(vertical="center", wrap_text=True)

        c_prop = ws1.cell(row=row_idx, column=4, value="")
        c_prop.fill = input_fill
        c_prop.font = Font(name="Calibri", size=11, bold=True, color="0F172A")
        c_prop.alignment = Alignment(vertical="center")

        priority_raw = str(item.get("review_priority", "high"))
        priority_label = "Yüksek" if priority_raw == "high" else ("Orta" if priority_raw == "medium" else "Düşük")
        c_prio = ws1.cell(row=row_idx, column=5, value=priority_label)
        if priority_label in priority_styles:
            fill_p, font_p = priority_styles[priority_label]
            c_prio.fill = fill_p
            c_prio.font = font_p
        c_prio.alignment = Alignment(horizontal="center", vertical="center")

        c_occ = ws1.cell(row=row_idx, column=6, value=item.get("occurrence_count", 0))
        c_occ.alignment = Alignment(horizontal="center", vertical="center")
        c_occ.font = Font(name="Calibri", size=10, bold=True)

        c_pg = ws1.cell(row=row_idx, column=7, value=_clean_str(_pages(item)))
        c_pg.alignment = Alignment(horizontal="center", vertical="center")
        c_pg.font = Font(name="Calibri", size=10)

        c_note = ws1.cell(row=row_idx, column=8, value="")
        c_note.alignment = Alignment(vertical="center")

        for col_i in range(1, 9):
            cell_target = ws1.cell(row=row_idx, column=col_i)
            cell_target.border = thin_border
            if is_even and col_i != 4 and col_i != 5:
                cell_target.fill = alt_fill

    ws1.freeze_panes = "A5"
    if missing:
        ws1.auto_filter.ref = f"A4:H{header_row_1 + len(missing)}"

    # ===================================================
    # SEKME 2: SÖZLÜKTE BULUNANLAR VE KISALTMALAR
    # ===================================================
    ws2 = wb.create_sheet(title="Sözlükte Bulunanlar")
    ws2.views.sheetView[0].showGridLines = True

    found_all = found + possible

    ws2.merge_cells("A1:G1")
    title2 = ws2["A1"]
    title2.value = _clean_str(f"📚 TBD BİLİŞİM SÖZLÜĞÜNDE BULUNAN TERİMLER VE KISALTMALAR")
    title2.font = Font(name="Calibri", size=14, bold=True, color="166534")
    title2.alignment = Alignment(vertical="center")
    ws2.row_dimensions[1].height = 28

    ws2.merge_cells("A2:G2")
    sub2 = ws2["A2"]
    sub2.value = _clean_str(f"Toplam Eşleşen Terim: {len(found_all)} Adet | TBD Sözlük Sürümü: {dict_version}")
    sub2.font = Font(name="Calibri", size=10, italic=True, color="475569")
    sub2.alignment = Alignment(vertical="center")
    ws2.row_dimensions[2].height = 20

    header_row_2 = 4
    ws2.row_dimensions[header_row_2].height = 28
    headers_2 = [
        ("No", 6),
        ("İngilizce Terim", 26),
        ("TBD Sözlük Türkçe Karşılığı", 34),
        ("Eşleşme Durumu", 26),
        ("Makaledeki Örnek Cümle (Bağlam)", 50),
        ("Geçiş", 10),
        ("Sayfalar", 14),
    ]

    header_fill_2 = PatternFill(start_color="166534", fill_type="solid")
    header_font_2 = Font(name="Calibri", size=11, bold=True, color="FFFFFF")

    for col_idx, (h_text, col_w) in enumerate(headers_2, 1):
        cell = ws2.cell(row=header_row_2, column=col_idx, value=h_text)
        cell.fill = header_fill_2
        cell.font = header_font_2
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        col_letter = get_column_letter(col_idx)
        ws2.column_dimensions[col_letter].width = col_w

    for idx, item in enumerate(found_all, 1):
        row_idx = header_row_2 + idx
        ws2.row_dimensions[row_idx].height = 32
        is_even = (idx % 2 == 0)

        translations = item.get("translations", [])
        if not translations:
            suggestions = item.get("possible_dictionary_terms", [])
            tr_text = " | ".join(str(v.get("tr", "")) for v in suggestions if isinstance(v, dict))
        else:
            tr_text = ", ".join(str(v) for v in translations)

        match_type = str(item.get("match_type", ""))
        if match_type == "singular_variant":
            status_desc = "Sözlükte Bulundu (Çoğul Eşleşme)"
        elif match_type == "exact":
            status_desc = "Sözlükte Kayıtlı"
        elif item.get("possible_dictionary_terms"):
            status_desc = "TBD Kısaltması"
        else:
            status_desc = "Sözlükte Bulundu"

        c_no = ws2.cell(row=row_idx, column=1, value=idx)
        c_no.alignment = Alignment(horizontal="center", vertical="center")
        c_no.font = Font(name="Calibri", size=10, color="64748B")

        c_term = ws2.cell(row=row_idx, column=2, value=_clean_str(item.get("term", "")))
        c_term.font = Font(name="Calibri", size=11, bold=True, color="0F172A")
        c_term.alignment = Alignment(vertical="center")

        c_tr = ws2.cell(row=row_idx, column=3, value=_clean_str(tr_text))
        c_tr.font = Font(name="Calibri", size=11, bold=True, color="166534")
        c_tr.alignment = Alignment(vertical="center")

        c_st = ws2.cell(row=row_idx, column=4, value=_clean_str(status_desc))
        c_st.font = Font(name="Calibri", size=10, color="334155")
        c_st.alignment = Alignment(horizontal="center", vertical="center")

        c_ctx = ws2.cell(row=row_idx, column=5, value=_clean_str(item.get("context", "")))
        c_ctx.font = Font(name="Calibri", size=10, italic=True, color="475569")
        c_ctx.alignment = Alignment(vertical="center", wrap_text=True)

        c_occ = ws2.cell(row=row_idx, column=6, value=item.get("occurrence_count", 0))
        c_occ.alignment = Alignment(horizontal="center", vertical="center")
        c_occ.font = Font(name="Calibri", size=10, bold=True)

        c_pg = ws2.cell(row=row_idx, column=7, value=_clean_str(_pages(item)))
        c_pg.alignment = Alignment(horizontal="center", vertical="center")
        c_pg.font = Font(name="Calibri", size=10)

        for col_i in range(1, 8):
            cell_target = ws2.cell(row=row_idx, column=col_i)
            cell_target.border = thin_border
            if is_even:
                cell_target.fill = alt_fill

    ws2.freeze_panes = "A5"
    if found_all:
        ws2.auto_filter.ref = f"A4:G{header_row_2 + len(found_all)}"

    wb.save(xlsx_path)


def _model_directory_name(model: object) -> str:
    """Model adını dosya sistemine uygun güvenli bir klasör adına çevirir."""
    value = unicodedata.normalize("NFKC", str(model or "bilinmeyen-model")).strip()
    value = re.sub(r"[\\/:*?\"<>|\s]+", "-", value)
    value = re.sub(r"[^A-Za-z0-9._-]+", "-", value).strip(".-_")
    return value.casefold() or "bilinmeyen-model"


def write_reports(result: dict[str, object], output_dir: Path) -> tuple[Path, Path]:
    """JSON, CSV ve 2 sekmeli gelişmiş XLSX raporlarını yazar."""
    output_dir = Path(output_dir)
    stem = Path(str(result["document"])).stem
    doc_dir = output_dir / _model_directory_name(result.get("model")) / stem
    doc_dir.mkdir(parents=True, exist_ok=True)
    json_path = doc_dir / "{}_terms.json".format(stem)
    csv_path = doc_dir / "{}_terim_raporu.csv".format(stem)
    xlsx_path = doc_dir / "{}_terim_raporu.xlsx".format(stem)

    json_path.write_text(
        json.dumps(result, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )
    rows = report_rows(result)
    fieldnames = [
        "İnceleme Durumu",
        "Eşleşme Türü",
        "Öncelik",
        "Önerilen İşlem",
        "İngilizce Terim",
        "Türkçe Karşılık",
        "Makaledeki Bağlam (Örnek Cümle)",
        "Önerilen Türkçe Karşılık (Komite)",
        "Yakın Sözlük Eşleşmesi",
        "Kanıt Sayfaları",
        "PDF'deki Geçiş Sayısı",
        "Açıklama",
    ]

    with csv_path.open("w", encoding="utf-8-sig", newline="") as file:
        writer = csv.DictWriter(file, fieldnames=fieldnames, delimiter=";")
        writer.writeheader()
        writer.writerows(rows)

    _export_styled_xlsx(result, xlsx_path)

    return json_path, csv_path
